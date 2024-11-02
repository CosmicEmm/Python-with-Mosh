import openpyxl as xl
#Chart: to add a chart, we need to import a couple of classes
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb ['Sheet1']

    for row in range(2, sheet.max_row + 1): #we don't need the first row as it only contains headings
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    #Create an instance of the Reference class to select a range of values for the chart
    values = Reference(sheet,
              min_row= 2,
              max_row= 4,
              min_col= 4,
              max_col= 4)

    #Create an instance of the BarChart class
    chart = BarChart()
    chart.add_data(values)
    #Add the chart to the spreadsheet
    sheet.add_chart(chart, 'f2')
    wb.save('transactions2.xlsx') #create a new file and makes the changes to that

process_workbook('transactions.xlsx')








